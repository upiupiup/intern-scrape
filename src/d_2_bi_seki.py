"""
d_2_bi_seki.py

Download arsip bulanan "Statistik Ekonomi dan Keuangan Indonesia" (SEKI)
dari situs Bank Indonesia (bi.go.id), ambil file TABEL8_1 dari setiap ZIP,
lalu parse data inflasi IHK Umum YoY menjadi CSV final.

Alur besar:
  1. Download / refresh sumber
     - Cari halaman rilis SEKI per bulan dari bulan terbaru mundur ke belakang.
     - Situs BI punya banyak pola URL lama/baru, jadi script mencoba banyak
       kandidat URL dan mendeteksi soft-404.
     - Dari halaman yang valid, script cari link ZIP publikasi.
     - ZIP tidak disimpan utuh; yang disimpan hanya file TABEL8_1
       (.xls/.xlsx/.pdf) ke:
       data/raw/groupD/seki_downloads/<TAHUN>/TABEL8_1_<BULAN>_<TAHUN>.<ext>
     - Kalau file TABEL8_1 lokal sudah ada, download/extract di-skip. Namun
       script tetap bisa mencari dan mencatat URL halaman rilis asli ke
       _edition_page_urls.csv kalau URL sumber belum tercatat, supaya kolom
       data_source di output tidak memakai URL tebakan.

  2. Parse data inflasi YoY
     - Baca semua file TABEL8_1 yang sudah tersimpan di folder raw.
     - Untuk .xls/.xlsx:
       cari sheet utama tabel 8.1, lalu ambil bagian:
       UMUM -> Tahun ke Tahun.
       Sheet historis dipakai sebagai gap-filler dengan prioritas lebih rendah.
     - Untuk .pdf:
       parser menyusun ulang teks dari karakter PDF, mencari baris
       "Tahun ke Tahun" / "Year on Year", lalu mencocokkan nilai ke kolom
       bulan berdasarkan posisi x agar nilai tidak bergeser saat ada kolom
       kosong/tanda "-".
     - Output periode diformat sebagai YYYY-MM.

  3. Gabung dan pilih sumber terbaik per periode
     - Satu periode bisa muncul di banyak edisi SEKI karena TABEL8_1 memuat
       data historis beberapa tahun.
     - Prioritas sumber:
       rank 1: sheet utama .xls/.xlsx
       rank 2: PDF
       rank 3: sheet historis .xls/.xlsx sebagai gap-filler
     - Dalam rank yang sama, edisi yang lebih baru menang karena publikasi BI
       yang lebih baru biasanya memuat revisi.

Output:
  - CSV final:
    data/processed/groupD/usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy.csv
  - Log parsing:
    data/processed/groupD/usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy_log.csv
  - Audit sumber per periode:
    data/processed/groupD/usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy_sumber_per_periode.csv
  - Log download:
    data/raw/groupD/seki_downloads/_log_download.csv
  - Mapping URL halaman edisi terverifikasi:
    data/raw/groupD/seki_downloads/_edition_page_urls.csv

Cara pakai:
    python d_2_bi_seki.py
    python d_2_bi_seki.py --start-year 2026 --start-month 7 --end-year 2004
    python d_2_bi_seki.py --stop-after-missing 6
    python d_2_bi_seki.py --pipeline-only
    python d_2_bi_seki.py --pipeline-only --no-secondary-sheets
"""

import argparse
import csv
import glob
import io
import os
import re
import sys
import time
import zipfile
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import math

import pandas as pd
import pdfplumber
import xlrd

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Konfigurasi
BULAN_ID = [
    "JANUARI",
    "FEBRUARI",
    "MARET",
    "APRIL",
    "MEI",
    "JUNI",
    "JULI",
    "AGUSTUS",
    "SEPTEMBER",
    "OKTOBER",
    "NOVEMBER",
    "DESEMBER",
]

BULAN_ABBR = {
    "JANUARI": "Jan",
    "FEBRUARI": "Feb",
    "MARET": "Mar",
    "APRIL": "April",
    "MEI": "Mei",
    "JUNI": "Juni",
    "JULI": "Juli",
    "AGUSTUS": "Agust",
    "SEPTEMBER": "Sept",
    "OKTOBER": "Okt",
    "NOVEMBER": "Nov",
    "DESEMBER": "Des",
}

SEKI_BASE = "https://www.bi.go.id/id/statistik/ekonomi-keuangan/seki/Pages/"
SEKI_BASE_LEGACY = "https://www.bi.go.id/id/statistik/seki/bulanan/Pages/"

OUTPUT_DIR = os.path.join("data", "raw", "groupD", "seki_downloads")
LOG_PATH = os.path.join(OUTPUT_DIR, "_log_download.csv")
EDITION_URL_LOG_PATH = os.path.join(OUTPUT_DIR, "_edition_page_urls.csv")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}

REQUEST_TIMEOUT = 30
SLEEP_BETWEEN_REQUESTS = 1.0
SLEEP_BETWEEN_CANDIDATES = 0.35
MAX_RETRIES = 3

SOFT_404_MARKERS = (
    "e404.aspx",
    "404 halaman tidak ditemukan",
    "halaman tidak ditemukan",
)

MANUAL_URL_OVERRIDE = {
    ("APRIL", 2005): f"{SEKI_BASE}seki-2005.zip.aspx",
    ("SEPTEMBER", 2004): f"{SEKI_BASE}SEP-2005.ZIP.aspx",
    ("JUNI", 2012): f"{SEKI_BASE}seki_0612-2.aspx",
    ("JULI", 2012): f"{SEKI_BASE}seki_0612-1.aspx",
}

BULAN_EN = {
    "JANUARI": "January",
    "FEBRUARI": "February",
    "MARET": "March",
    "APRIL": "April",
    "MEI": "May",
    "JUNI": "June",
    "JULI": "July",
    "AGUSTUS": "August",
    "SEPTEMBER": "September",
    "OKTOBER": "October",
    "NOVEMBER": "November",
    "DESEMBER": "December",
}


# Helper functions
def fetch_with_retry(session, url, max_retries=MAX_RETRIES, **kwargs):
    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, **kwargs)
            return resp
        except requests.RequestException as e:
            print(f"    [retry {attempt}/{max_retries}] error: {e}")
            time.sleep(2 * attempt)
    return None


def is_soft_404(resp):
    if resp is None:
        return True
    final_url = (resp.url or "").lower()
    if "e404.aspx" in final_url:
        return True
    head = (resp.text or "")[:4000].lower()
    for marker in SOFT_404_MARKERS:
        if marker in head:
            return True
    return False


def build_page_url_candidates(bulan, tahun):
    bulan_title = bulan.capitalize()
    bulan_abbr = BULAN_ABBR.get(bulan, bulan_title)
    bulan_abbr_lower = bulan_abbr.lower()
    bulan_lower = bulan.lower()
    tahun_2digit = f"{tahun % 100:02d}"
    bulan_num = BULAN_ID.index(bulan) + 1
    mm = f"{bulan_num:02d}"
    yy = f"{tahun % 100:02d}"
    bulan_abbr3 = bulan_title[:3]

    candidates = [
        f"{SEKI_BASE}SEKI-{bulan}-{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan_title}-{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan_abbr}-{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan_abbr_lower}-{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan}{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan_title}{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{bulan_abbr}{tahun}.aspx",
        f"{SEKI_BASE}SEKI%20{bulan_abbr}%20{tahun}.aspx",
        f"{SEKI_BASE}SEKI%20{bulan_title}%20{tahun}.aspx",
        f"{SEKI_BASE}SEKI%20{bulan}%20{tahun}.aspx",
        f"{SEKI_BASE}{bulan_abbr_lower}-{tahun}.zip.aspx",
        f"{SEKI_BASE}{bulan_abbr_lower}-{tahun}.aspx",
        f"{SEKI_BASE}{bulan_lower}-{tahun}.zip.aspx",
        f"{SEKI_BASE}{bulan_lower}-{tahun}.aspx",
        f"{SEKI_BASE_LEGACY}SEKI-{bulan}-{tahun}.aspx",
        f"{SEKI_BASE}seki_{bulan_num:02d}{tahun_2digit}.aspx",
        f"{SEKI_BASE}seki_{mm}{yy}.aspx",
        f"{SEKI_BASE}seki_{mm}{yy}-1.aspx",
        f"{SEKI_BASE}seki_{mm}{yy}-2.aspx",
        f"{SEKI_BASE}seki_{mm}{yy}_rev.aspx",
        f"{SEKI_BASE}seki-{tahun}{mm}.aspx",
        f"{SEKI_BASE}SEKI-{bulan}{tahun}-tahap1.aspx",
        f"{SEKI_BASE}SEKI-{bulan_title}{tahun}-tahap1.aspx",
        f"{SEKI_BASE}SEKI_{bulan}_{tahun}.aspx",
        f"{SEKI_BASE}SEKI---{bulan_title}-{tahun}.aspx",
        f"{SEKI_BASE}SEKI-{BULAN_EN[bulan]}-{tahun}.aspx",
        f"{SEKI_BASE}{bulan_lower}{yy}.aspx",
        f"{SEKI_BASE}{bulan_abbr3.lower()}-{yy}.aspx",
        f"{SEKI_BASE}{bulan_lower}-{yy}.aspx",
        f"{SEKI_BASE}seki%20{bulan_abbr3.lower()}%20{yy}.aspx",
        f"{SEKI_BASE}seki%20{mm}{yy}.aspx",
        f"{SEKI_BASE}{bulan_title}%20-%20{tahun}.aspx",
        f"{SEKI_BASE}SEKI%20{bulan_abbr3}%20{tahun}.aspx",
        f"{SEKI_BASE}{bulan_lower}%20{tahun}.aspx",
    ]

    seen = set()
    unique_candidates = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            unique_candidates.append(c)
    return unique_candidates


def find_page_url(session, bulan, tahun):
    override_url = MANUAL_URL_OVERRIDE.get((bulan, tahun))
    if override_url:
        resp = fetch_with_retry(session, override_url)
        if resp is not None and resp.status_code == 200 and not is_soft_404(resp):
            return override_url, resp.text
        return None, None

    for url in build_page_url_candidates(bulan, tahun):
        resp = fetch_with_retry(session, url)
        if resp is not None and resp.status_code == 200 and not is_soft_404(resp):
            return url, resp.text
        time.sleep(SLEEP_BETWEEN_CANDIDATES)
    return None, None


def find_zip_link(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.lower().endswith(".zip"):
            if href.startswith("http"):
                return href
            if href.startswith("/"):
                return "https://www.bi.go.id" + href
            return requests.compat.urljoin(base_url, href)

    match = re.search(r'href=["\']([^"\']+\.zip)["\']', html, re.IGNORECASE)
    if match:
        href = match.group(1)
        if href.startswith("http"):
            return href
        return requests.compat.urljoin(base_url, href)

    return None


TABEL81_PATTERN = re.compile(r"^tabel8_1\.(xls|xlsx|pdf)$", re.IGNORECASE)


def find_tabel81_member(zf):
    for name in zf.namelist():
        base = os.path.basename(name)
        if TABEL81_PATTERN.match(base):
            return name
    return None


def download_and_extract_tabel81(session, zip_url, dest_path_no_ext):
    resp = fetch_with_retry(session, zip_url, stream=False)
    if resp is None or resp.status_code != 200:
        return None
    content_type = resp.headers.get("Content-Type", "")
    if "html" in content_type.lower():
        return None

    try:
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
    except zipfile.BadZipFile:
        return None

    member = find_tabel81_member(zf)
    if member is None:
        return None

    ext = os.path.splitext(member)[1].lower()
    dest_path = dest_path_no_ext + ext
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with zf.open(member) as src, open(dest_path, "wb") as out:
        out.write(src.read())
    return dest_path


def record_edition_page_url(tahun, bulan, page_url):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_header = not os.path.exists(EDITION_URL_LOG_PATH)
    with open(EDITION_URL_LOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(["tahun", "bulan", "page_url"])
        writer.writerow([tahun, bulan, page_url])


def load_edition_page_urls(raw_dir=OUTPUT_DIR):
    path = os.path.join(raw_dir, "_edition_page_urls.csv")
    url_map = {}
    if not os.path.exists(path):
        return url_map
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                tahun = int(row["tahun"])
            except (KeyError, ValueError):
                continue
            bulan = (row.get("bulan") or "").strip().upper()
            url = row.get("page_url") or ""
            if bulan and url:
                url_map[(bulan, tahun)] = url
    return url_map


def log_row(rows_written_header, tahun, bulan, status, detail):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_header = not os.path.exists(LOG_PATH)
    with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(["tahun", "bulan", "status", "detail", "timestamp"])
        writer.writerow(
            [tahun, bulan, status, detail, datetime.now().isoformat(timespec="seconds")]
        )


# Main
def month_iter(start_year, start_month, end_year, end_month=1):
    y, m = start_year, start_month
    while (y > end_year) or (y == end_year and m >= end_month):
        yield y, m
        m -= 1
        if m == 0:
            m = 12
            y -= 1


def main():
    parser = argparse.ArgumentParser(
        description="Download arsip ZIP bulanan SEKI dari bi.go.id"
    )
    now = datetime.now()
    parser.add_argument("--start-year", type=int, default=now.year)
    parser.add_argument("--start-month", type=int, default=now.month)
    parser.add_argument("--end-year", type=int, default=2004)
    parser.add_argument("--end-month", type=int, default=1)
    parser.add_argument(
        "--stop-after-missing",
        type=int,
        default=0,
        help="Stop kalau N bulan berturut-turut sama sekali gak ketemu (halaman "
        "maupun zip). 0 = jangan pernah stop otomatis (default).",
    )
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        default=True,
        help="Skip download kalau file zip sudah ada (default: True).",
    )
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    session = requests.Session()
    edition_url_map = load_edition_page_urls(OUTPUT_DIR)

    consecutive_missing = 0
    total_ok, total_missing = 0, 0

    for tahun, m_idx in month_iter(
        args.start_year, args.start_month, args.end_year, args.end_month
    ):
        bulan = BULAN_ID[m_idx - 1]
        label = f"{bulan} {tahun}"
        print(f"[{label}] mencari halaman rilis ...")

        dest_dir = os.path.join(OUTPUT_DIR, str(tahun))
        dest_base = os.path.join(dest_dir, f"TABEL8_1_{bulan}_{tahun}")

        existing_matches = glob.glob(dest_base + ".*")
        if args.skip_existing and existing_matches:
            print(f"    sudah ada, skip: {existing_matches[0]}")
            if (bulan, tahun) not in edition_url_map:
                page_url, _html = find_page_url(session, bulan, tahun)
                if page_url:
                    record_edition_page_url(tahun, bulan, page_url)
                    edition_url_map[(bulan, tahun)] = page_url
                    print(f"    URL sumber dicatat: {page_url}")
                else:
                    print("    URL sumber belum ketemu (data lokal tetap dipakai)")
                time.sleep(SLEEP_BETWEEN_REQUESTS)
            log_row(True, tahun, bulan, "skipped_existing", existing_matches[0])
            consecutive_missing = 0
            total_ok += 1
            continue

        page_url, html = find_page_url(session, bulan, tahun)

        if page_url is None:
            print(f"    halaman TIDAK ditemukan untuk {label} (semua pola URL dicoba)")
            log_row(True, tahun, bulan, "page_not_found", "")
            consecutive_missing += 1
            total_missing += 1
        else:
            zip_url = find_zip_link(html, page_url)
            if zip_url is None:
                print(f"    halaman ketemu ({page_url}) tapi TIDAK ada link .zip")
                log_row(True, tahun, bulan, "page_found_no_zip", page_url)
                consecutive_missing += 1
                total_missing += 1
            else:
                print(f"    halaman: {page_url}")
                print(f"    zip ketemu: {zip_url}")
                record_edition_page_url(tahun, bulan, page_url)
                saved_path = download_and_extract_tabel81(session, zip_url, dest_base)
                if saved_path:
                    print(f"    TABEL8_1 ketemu, disimpan -> {saved_path}")
                    log_row(True, tahun, bulan, "downloaded", saved_path)
                    consecutive_missing = 0
                    total_ok += 1
                else:
                    print(f"    zip ketemu tapi TIDAK ada TABEL8_1 di dalamnya, skip")
                    log_row(True, tahun, bulan, "tabel81_not_in_zip", zip_url)
                    consecutive_missing = 0
                    total_missing += 1

        if args.stop_after_missing and consecutive_missing >= args.stop_after_missing:
            print(
                f"\nBerhenti: {consecutive_missing} bulan berturut-turut tidak "
                f"ditemukan (berhenti di {label})."
            )
            break

        time.sleep(SLEEP_BETWEEN_REQUESTS)

    print(f"\nSelesai. Berhasil: {total_ok}, tidak ketemu/gagal: {total_missing}")
    print(f"Log lengkap ada di: {LOG_PATH}")


#   PIPELINE TAMBAHAN untuk usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy
PROCESSED_DIR = os.path.join("data", "processed", "groupD")
PROCESSED_BASENAME = "usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy"
PROCESSED_CSV_PATH = os.path.join(PROCESSED_DIR, f"{PROCESSED_BASENAME}.csv")
PIPELINE_LOG_PATH = os.path.join(PROCESSED_DIR, f"{PROCESSED_BASENAME}_log.csv")
SOURCE_AUDIT_PATH = os.path.join(
    PROCESSED_DIR, f"{PROCESSED_BASENAME}_sumber_per_periode.csv"
)

KATEGORI_LABEL = "Umum"
SATUAN_LABEL = "persen"

MONTH_ALIASES = {
    "jan": 1,
    "januari": 1,
    "january": 1,
    "feb": 2,
    "februari": 2,
    "february": 2,
    "mar": 3,
    "maret": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "mei": 5,
    "may": 5,
    "jun": 6,
    "juni": 6,
    "june": 6,
    "jul": 7,
    "juli": 7,
    "july": 7,
    "agu": 8,
    "ags": 8,
    "agt": 8,
    "agust": 8,
    "agustus": 8,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "okt": 10,
    "oktober": 10,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "des": 12,
    "desember": 12,
    "dec": 12,
    "december": 12,
}

FLOAT_RE = re.compile(r"^-?\d+\.\d+$")
TABEL81_FILE_RE = re.compile(
    r"^TABEL8_1_([A-Za-z]+)_(\d{4})\.(xls|xlsx|pdf)$", re.IGNORECASE
)
HIST_SHEET_RE = re.compile(r"^Th\s*\d{4}\s*-\s*\d{4}$", re.IGNORECASE)


def _month_from_token(tok):
    if isinstance(tok, datetime):
        return tok.month
    t = re.sub(r"[^A-Za-z]", "", str(tok)).strip().lower()
    return MONTH_ALIASES.get(t)


def _month_shift(year, month, k):
    idx = year * 12 + (month - 1) + k
    return idx // 12, idx % 12 + 1


def edition_source_url(bulan, tahun, url_map=None):
    if url_map:
        real = url_map.get((bulan, tahun))
        if real:
            return real, True
    override = MANUAL_URL_OVERRIDE.get((bulan, tahun))
    if override:
        return override, True
    candidates = build_page_url_candidates(bulan, tahun)
    return (candidates[0] if candidates else ""), False


def discover_tabel81_files(raw_dir=OUTPUT_DIR):
    found = []
    pattern = os.path.join(raw_dir, "*", "TABEL8_1_*.*")
    for path in glob.glob(pattern):
        fname = os.path.basename(path)
        m = TABEL81_FILE_RE.match(fname)
        if not m:
            continue
        bulan_raw, tahun_raw, ext = m.groups()
        bulan = bulan_raw.upper()
        if bulan not in BULAN_ID:
            continue
        found.append(
            {
                "bulan": bulan,
                "tahun": int(tahun_raw),
                "path": path,
                "ext": ext.lower(),
            }
        )
    found.sort(key=lambda r: (r["tahun"], BULAN_ID.index(r["bulan"])))
    return found


class _XlsSheetAdapter:
    def __init__(self, sheet, datemode):
        self._sheet = sheet
        self._datemode = datemode
        self.nrows = sheet.nrows
        self.ncols = sheet.ncols

    def cell_value(self, r, c):
        v = self._sheet.cell_value(r, c)
        try:
            ctype = self._sheet.cell_type(r, c)
        except Exception:
            return v
        if ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(v, self._datemode)
            except Exception:
                return v
        return v


class _XlsxSheetAdapter:
    def __init__(self, ws):
        self._ws = ws
        self.nrows = ws.max_row or 0
        self.ncols = ws.max_column or 0

    def cell_value(self, r, c):
        v = self._ws.cell(row=r + 1, column=c + 1).value
        return v if v is not None else ""


def _find_label_row(
    sheet, target_labels_exact, search_cols=range(0, 6), row_range=None
):
    row_range = row_range if row_range is not None else range(sheet.nrows)
    for r in row_range:
        for c in search_cols:
            if c >= sheet.ncols:
                continue
            v = sheet.cell_value(r, c)
            if isinstance(v, str) and v.strip().lower() in target_labels_exact:
                return r, c
    return None, None


def _find_month_header_row(sheet, max_row=20):
    best_row, best_count = None, 0
    for r in range(min(max_row, sheet.nrows)):
        cnt = sum(
            1
            for c in range(sheet.ncols)
            if isinstance(sheet.cell_value(r, c), (str, datetime))
            and _month_from_token(sheet.cell_value(r, c)) is not None
        )
        if cnt > best_count:
            best_count, best_row = cnt, r
    return best_row, best_count


def parse_umum_yoy_from_sheet(sheet):
    umum_row, _ = _find_label_row(sheet, {"umum"})
    if umum_row is None:
        return None, "tidak ada label 'UMUM'"

    yoy_row, _ = _find_label_row(
        sheet,
        {"tahun ke tahun"},
        row_range=range(umum_row, min(umum_row + 10, sheet.nrows)),
    )
    if yoy_row is None:
        return None, "ada 'UMUM' tapi tidak ada baris 'Tahun Ke Tahun' di dekatnya"

    month_row, month_count = _find_month_header_row(sheet, max_row=20)
    if month_row is None or month_count < 1:
        return None, f"header baris bulan tidak terdeteksi (match={month_count})"
    indeks_row, _ = _find_label_row(
        sheet, {"indeks"}, row_range=range(max(0, umum_row - 2), yoy_row + 1)
    )

    results = []
    current_year = None
    last_month = None
    for c in range(sheet.ncols):
        mv = sheet.cell_value(month_row, c)
        m = _month_from_token(mv) if isinstance(mv, (str, datetime)) else None
        if m is None:
            continue
        if isinstance(mv, datetime):
            current_year = mv.year
        else:
            if current_year is None:
                for rr in range(max(0, month_row - 3), month_row):
                    yv = sheet.cell_value(rr, c)
                    if (
                        isinstance(yv, (int, float))
                        and 1990 <= yv <= 2100
                        and float(yv).is_integer()
                    ):
                        current_year = int(yv)
                        break
                if current_year is None:
                    year_row = month_row - 1 if month_row > 0 else month_row
                    for cc in range(0, c + 1):
                        yv = sheet.cell_value(year_row, cc)
                        if (
                            isinstance(yv, (int, float))
                            and 1990 <= yv <= 2100
                            and float(yv).is_integer()
                        ):
                            current_year = int(yv)
            if current_year is None:
                continue
            if last_month is not None and m <= last_month:
                current_year += 1
        last_month = m
        if indeks_row is not None:
            idx_val = sheet.cell_value(indeks_row, c)
            if not (isinstance(idx_val, (int, float)) and 0 < idx_val < 1000):
                continue
        val = sheet.cell_value(yoy_row, c)
        if isinstance(val, (int, float)):
            results.append((current_year * 100 + m, float(val)))
    if not results:
        return (
            None,
            "header bulan ketemu tapi tidak ada nilai numerik yang bisa dipasangkan",
        )
    return results, None


def parse_xls_file(path):
    warnings = []
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".xls":
            wb = xlrd.open_workbook(path)
            sheet_names = list(wb.sheet_names())
            get_sheet = lambda n: _XlsSheetAdapter(wb.sheet_by_name(n), wb.datemode)
        elif ext == ".xlsx":
            if openpyxl is None:
                return [], {}, ["file .xlsx ketemu tapi openpyxl tidak terinstall"]
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet_names = list(wb.sheetnames)
            get_sheet = lambda n: _XlsxSheetAdapter(wb[n])
        else:
            return [], {}, [f"ekstensi tidak didukung: {ext}"]
    except Exception as e:
        return [], {}, [f"gagal membuka workbook: {e}"]

    non_hist = [n for n in sheet_names if not HIST_SHEET_RE.match(n.strip())]
    primary_name = None
    for n in non_hist:
        norm = re.sub(r"[\s_]", "", n).lower()
        if norm.startswith("8.1") or norm.startswith("81"):
            primary_name = n
            break
    if primary_name is None:
        primary_name = (
            non_hist[-1] if non_hist else (sheet_names[-1] if sheet_names else None)
        )

    primary_records = []
    secondary_records = {}
    for name in sheet_names:
        try:
            sh = get_sheet(name)
        except Exception as e:
            warnings.append(f"sheet '{name}': gagal dibuka ({e})")
            continue
        recs, err = parse_umum_yoy_from_sheet(sh)
        if err:
            if name == primary_name:
                warnings.append(f"sheet utama '{name}': {err}")
            continue
        if name == primary_name:
            primary_records = recs
        else:
            secondary_records[name] = recs

    if primary_name is None:
        warnings.append("tidak ada sheet sama sekali di workbook ini")
    elif not primary_records and not any(
        f"sheet utama '{primary_name}'" in w for w in warnings
    ):
        warnings.append(f"sheet utama '{primary_name}' tidak menghasilkan data")

    return primary_records, secondary_records, warnings


# Parser PDF Halaman ID ('Tahun ke Tahun') + halaman EN ('Year on Year').
def _clean_lines(page, gap_thresh=3.0, line_tol=2.0):
    chars = [c for c in page.chars if not c["text"].isspace()]
    chars.sort(key=lambda c: (c["top"], c["x0"]))

    raw_lines = []
    cur_line, cur_top = [], None

    for c in chars:
        if cur_top is None:
            cur_top = c["top"]
        elif abs(c["top"] - cur_top) > line_tol:
            if cur_line:
                raw_lines.append(cur_line)
            cur_line = []
            cur_top = c["top"]
        cur_line.append(c)
    if cur_line:
        raw_lines.append(cur_line)

    lines = []
    for raw in raw_lines:
        raw.sort(key=lambda c: c["x0"])
        tokens = []
        cur_word, cur_x0, last_x1 = "", None, None

        def flush_word():
            nonlocal cur_word, cur_x0
            if cur_word:
                tokens.append((cur_x0, cur_word))
            cur_word, cur_x0 = "", None

        for c in raw:
            if last_x1 is not None and c["x0"] - last_x1 > gap_thresh:
                flush_word()
            if cur_x0 is None:
                cur_x0 = c["x0"]
            cur_word += c["text"]
            last_x1 = max(last_x1, c["x1"]) if last_x1 is not None else c["x1"]
        flush_word()
        if tokens:
            lines.append(tokens)
    return lines


def _find_target_line(lines, labels):
    for tokens in lines:
        joined = "".join(t for _, t in tokens).lower()
        if any(lab in joined for lab in labels):
            return tokens
    return None


def _extract_float_values(tokens):
    return [float(t) for _, t in tokens if FLOAT_RE.match(t)]


def _extract_value_slots(tokens):
    slots = []
    for x0, t in tokens:
        if FLOAT_RE.match(t):
            slots.append((x0, float(t)))
        elif t == "-":
            slots.append((x0, None))
    slots.sort(key=lambda s: s[0])
    return slots


def _align_values_to_months(month_cols, value_slots, x_tol=12.0):
    used = [False] * len(value_slots)
    out = []
    for mx in month_cols:
        best_i, best_d = None, None
        for i, (vx, _v) in enumerate(value_slots):
            if used[i]:
                continue
            d = abs(vx - mx)
            if d <= x_tol and (best_d is None or d < best_d):
                best_i, best_d = i, d
        if best_i is None:
            out.append(None)
        else:
            used[best_i] = True
            out.append(value_slots[best_i][1])
    return out


def _find_month_header(lines, limit=25):
    best_count, best_tokens = 0, None
    for tokens in lines[:limit]:
        cnt = sum(1 for _, t in tokens if _month_from_token(t) is not None)
        if cnt > best_count:
            best_count, best_tokens = cnt, tokens
    month_cols = (
        [x0 for x0, t in best_tokens if _month_from_token(t) is not None]
        if best_tokens
        else []
    )
    return best_tokens, best_count, month_cols


def parse_pdf_file(path, edition_bulan, edition_tahun):
    edition_month_num = BULAN_ID.index(edition_bulan) + 1
    label_sets = (("tahunketahun",), ("yearonyear",))
    warnings = []
    page_infos = []

    try:
        with pdfplumber.open(path) as pdf:
            if len(pdf.pages) == 0:
                return [], ["PDF tidak punya halaman"]
            for pi, page in enumerate(pdf.pages):
                if not page.chars:
                    warnings.append(
                        f"halaman {pi + 1}: tidak ada layer teks sama sekali "
                        f"(kemungkinan hasil scan gambar, butuh OCR gambar terpisah) - dilewati"
                    )
                    continue
                lines = _clean_lines(page)
                target = None
                for labs in label_sets:
                    target = _find_target_line(lines, labs)
                    if target:
                        break
                if target is None:
                    continue
                month_tokens, month_count, month_cols = _find_month_header(lines)
                if month_count == 0:
                    warnings.append(
                        f"halaman {pi + 1}: baris 'Tahun Ke Tahun'/'Year on Year' ketemu, "
                        f"tapi header nama bulan tidak terdeteksi - dilewati"
                    )
                    continue
                value_slots = _extract_value_slots(target)
                monthly_values = _align_values_to_months(month_cols, value_slots)
                n_missing = sum(1 for v in monthly_values if v is None)
                if n_missing:
                    warnings.append(
                        f"halaman {pi + 1}: {n_missing} dari {month_count} kolom bulan "
                        f"tidak punya nilai (tanda '-' di tabel asli, atau tidak "
                        f"terbaca) - kolom lain tetap dipakai apa adanya"
                    )
                page_infos.append(
                    {
                        "page": pi + 1,
                        "month_count": month_count,
                        "monthly_values": monthly_values,
                    }
                )
    except Exception as e:
        return [], [f"gagal membaca PDF: {e}"]

    if not page_infos:
        warnings.append(
            "tidak ada baris 'Tahun ke Tahun' / 'Year on Year' ditemukan di PDF ini"
        )
        return [], warnings

    total_months = sum(p["month_count"] for p in page_infos)
    periods = [
        _month_shift(edition_tahun, edition_month_num, -(total_months - 1 - i))
        for i in range(total_months)
    ]

    idx = 0
    records = []
    for p in page_infos:
        n = p["month_count"]
        p_periods = periods[idx : idx + n]
        idx += n
        vals = p["monthly_values"]
        for (y, m), v in zip(p_periods, vals):
            if v is None:
                continue
            records.append((y * 100 + m, v))
    return records, warnings


# Orkestrasi pipeline
def run_pipeline(
    raw_dir=OUTPUT_DIR, processed_dir=PROCESSED_DIR, include_secondary_sheets=True
):
    os.makedirs(processed_dir, exist_ok=True)
    processed_csv_path = os.path.join(processed_dir, f"{PROCESSED_BASENAME}.csv")
    pipeline_log_path = os.path.join(processed_dir, f"{PROCESSED_BASENAME}_log.csv")
    source_audit_path = os.path.join(
        processed_dir, f"{PROCESSED_BASENAME}_sumber_per_periode.csv"
    )
    files = discover_tabel81_files(raw_dir)
    url_map = load_edition_page_urls(raw_dir)
    if url_map:
        print(
            f"Ditemukan {len(url_map)} URL edisi terverifikasi di "
            f"_edition_page_urls.csv (dipakai duluan utk kolom data_source)."
        )
    else:
        print(
            "Tidak ada _edition_page_urls.csv -- kolom data_source akan "
            "pakai TEBAKAN kandidat pertama (bisa saja 404, tidak dijamin "
            "valid). Jalankan proses download di sesi ini dulu kalau mau "
            "link yang terverifikasi."
        )

    if not files:
        print(
            f"Tidak ada file TABEL8_1 ditemukan di {raw_dir} — jalankan dulu proses download di atas."
        )

    best = {}

    def consider(periode, value, rank, tahun, bulan, source_url, asal, verified):
        edition_key = (tahun, BULAN_ID.index(bulan))
        cur = best.get(periode)
        better = (
            cur is None
            or rank < cur["rank"]
            or (rank == cur["rank"] and edition_key > cur["edition_key"])
        )
        if better:
            best[periode] = {
                "value": value,
                "rank": rank,
                "edition_key": edition_key,
                "tahun": tahun,
                "bulan": bulan,
                "source_url": source_url,
                "asal": asal,
                "verified": verified,
            }

    log_entries = []
    print(f"Ditemukan {len(files)} file TABEL8_1 di {raw_dir}")

    for f in files:
        bulan, tahun, path, ext = f["bulan"], f["tahun"], f["path"], f["ext"]
        label = f"{bulan} {tahun}"
        source_url, url_verified = edition_source_url(bulan, tahun, url_map)
        status, detail, n_pts = "ok", "", 0

        if ext in ("xls", "xlsx"):
            primary_records, secondary_records, warnings = parse_xls_file(path)
            for periode, value in primary_records:
                consider(
                    periode,
                    value,
                    1,
                    tahun,
                    bulan,
                    source_url,
                    "sheet utama (8.1)",
                    url_verified,
                )
            n_pts += len(primary_records)
            if include_secondary_sheets:
                for sheet_name, recs in secondary_records.items():
                    for periode, value in recs:
                        if value == 0.0:
                            continue
                        consider(
                            periode,
                            value,
                            3,
                            tahun,
                            bulan,
                            source_url,
                            f"sheet historis '{sheet_name}'",
                            url_verified,
                        )
                    n_pts += len(recs)
            if warnings:
                status = (
                    "partial"
                    if (primary_records or any(secondary_records.values()))
                    else "failed"
                )
                detail = " | ".join(warnings)
        elif ext == "pdf":
            records, warnings = parse_pdf_file(path, bulan, tahun)
            for periode, value in records:
                consider(
                    periode, value, 2, tahun, bulan, source_url, "pdf", url_verified
                )
            n_pts = len(records)
            if warnings:
                status = "partial" if records else "failed"
                detail = " | ".join(warnings)
        else:
            status, detail = "failed", f"ekstensi tidak dikenal: {ext}"

        log_entries.append(
            {
                "tahun": tahun,
                "bulan": bulan,
                "file": path,
                "ext": ext,
                "status": status,
                "n_titik_bulan_terbaca": n_pts,
                "detail": detail,
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            }
        )
        flag = "" if status == "ok" else f"  <-- {status.upper()}: {detail}"
        print(f"  [{label}] {status} ({n_pts} titik){flag}")

    periods_sorted = sorted(best.keys())
    rows = []
    for periode in periods_sorted:
        info = best[periode]
        rows.append(
            {
                "id": None,
                "periode": periode,
                "inflasi_yoy_persen": info["value"],
                "kategori": KATEGORI_LABEL,
                "satuan": SATUAN_LABEL,
                "data_source": info["source_url"],
                "_edisi_sumber": f"{info['bulan']} {info['tahun']}",
                "_asal": info["asal"],
                "_data_source_verified": info["verified"],
            }
        )
    for i, r in enumerate(rows, start=1):
        r["id"] = i

    df = pd.DataFrame(
        rows,
        columns=[
            "id",
            "periode",
            "inflasi_yoy_persen",
            "kategori",
            "satuan",
            "data_source",
            "_edisi_sumber",
            "_asal",
            "_data_source_verified",
        ],
    )
    df["periode"] = df["periode"].apply(
        lambda p: f"{int(p) // 100:04d}-{int(p) % 100:02d}"
    )
    df_final = df[
        ["id", "periode", "inflasi_yoy_persen", "kategori", "satuan", "data_source"]
    ]
    df_final.to_csv(processed_csv_path, index=False, encoding="utf-8-sig")

    log_df = pd.DataFrame(
        log_entries,
        columns=[
            "tahun",
            "bulan",
            "file",
            "ext",
            "status",
            "n_titik_bulan_terbaca",
            "detail",
            "timestamp",
        ],
    )
    log_df.to_csv(pipeline_log_path, index=False, encoding="utf-8-sig")

    df[
        [
            "periode",
            "inflasi_yoy_persen",
            "_edisi_sumber",
            "_asal",
            "_data_source_verified",
        ]
    ].to_csv(source_audit_path, index=False, encoding="utf-8-sig")
    n_unverified = int((~df["_data_source_verified"]).sum()) if not df.empty else 0
    if n_unverified:
        print(
            f"PERINGATAN: {n_unverified} baris pakai data_source hasil TEBAKAN "
            f"(bukan URL terverifikasi dari _edition_page_urls.csv) -- link ini "
            f"tidak dijamin valid, cek kolom _data_source_verified di "
            f"{source_audit_path}."
        )

    n_ok = sum(1 for r in log_entries if r["status"] == "ok")
    n_partial = sum(1 for r in log_entries if r["status"] == "partial")
    n_failed = sum(1 for r in log_entries if r["status"] == "failed")
    print(
        f"\nRingkasan proses: {n_ok} file ok, {n_partial} partial, {n_failed} failed (detail di log)."
    )
    if not df_final.empty:
        print(
            f"Output final: {processed_csv_path} "
            f"({len(df_final)} baris, periode {df_final['periode'].min()}..{df_final['periode'].max()})"
        )
    else:
        print(
            f"Output final: {processed_csv_path} (0 baris — cek log, kemungkinan belum ada file mentah)"
        )
    print(f"Log proses: {pipeline_log_path}")
    print(f"Audit sumber per periode (bonus): {source_audit_path}")
    return df_final, log_df


def build_pipeline_argparser():
    p = argparse.ArgumentParser(
        description="Proses TABEL8_1 (xls/xlsx/pdf) yang sudah didownload -> "
        "usecase_ekonomi.ekonomi_tren_inflasi_bulanan_yoy.csv"
    )
    p.add_argument("--raw-dir", default=OUTPUT_DIR)
    p.add_argument("--processed-dir", default=PROCESSED_DIR)
    p.add_argument(
        "--no-secondary-sheets",
        action="store_true",
        help="Kalau diset, sheet historis ('Th <awal>-<akhir>') di dalam xls "
        "TIDAK dipakai sebagai gap-filler.",
    )
    return p


def run_pipeline_cli(argv=None):
    args = build_pipeline_argparser().parse_args(argv)
    return run_pipeline(
        raw_dir=args.raw_dir,
        processed_dir=args.processed_dir,
        include_secondary_sheets=not args.no_secondary_sheets,
    )


if __name__ == "__main__":
    if "--pipeline-only" in sys.argv:
        sys.argv.remove("--pipeline-only")
        run_pipeline_cli()
    else:
        main()
        run_pipeline_cli([])
