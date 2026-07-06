"""
Extract Grup D (Bank Indonesia) tables into clean CSV files.

Input yang didukung:
1) TABEL8_1.pdf dari SEKI BI -> inflasi bulanan YoY dari baris UMUM / Tahun Ke Tahun.
2) SUSPI zip/xlsx -> sheet "4. Total Public Sector" saja, file IDR saja, periode 2019-2025.

Default script ini auto-download sumber dari web BI:
- https://www.bi.go.id/SEKI/tabel/TABEL8_1.pdf
- halaman SUSPI per triwulan: https://www.bi.go.id/id/statistik/ekonomi-keuangan/suspi/Pages/SUSPI_TWI_2019.aspx, dst.
  Dari setiap halaman SUSPI, script cari lampiran zip tabel SUSPI, lalu ambil workbook IDR saja.

Output filename mengikuti skema:
  <No>_<nama_tabel_db_tanpa_schema>.csv
Contoh:
  2_ekonomi_tren_inflasi_bulanan_yoy.csv
  4_ekonomi_tren_hutang_bulanan.csv

Contoh run:
  python src/extract_groupD_bi_tables.py --input-dir "data/raw/groupD" --schema "data/raw/aufi_Copy of Skema Data Ekonomi.xlsm" --output-dir "data/raw/groupD"

Kalau tidak mau download ulang dan hanya pakai file lokal:
  python src/extract_groupD_bi_tables.py --input-dir "data/raw/groupD" --schema "data/raw/aufi_Copy of Skema Data Ekonomi.xlsm" --output-dir "data/raw/groupD" --no-download

Install deps:
  pip install pandas openpyxl pdfplumber
"""

from __future__ import annotations

import argparse
import io
import os
import re
import sys
import time
import zipfile
import unicodedata
from html import unescape
from urllib.error import HTTPError, URLError
from urllib.parse import unquote, urljoin, urlparse
from urllib.request import Request, urlopen
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


SEKI_TABEL8_1_URL = "https://www.bi.go.id/SEKI/tabel/TABEL8_1.pdf"
SUSPI_BASE_URL = "https://www.bi.go.id/id/statistik/ekonomi-keuangan/suspi"
SUSPI_PAGES_URL = f"{SUSPI_BASE_URL}/Pages"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)


MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "januari": 1,
    "feb": 2,
    "february": 2,
    "februari": 2,
    "mar": 3,
    "march": 3,
    "maret": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "mei": 5,
    "jun": 6,
    "june": 6,
    "juni": 6,
    "jul": 7,
    "july": 7,
    "juli": 7,
    "aug": 8,
    "august": 8,
    "agustus": 8,
    "ags": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "okt": 10,
    "october": 10,
    "oktober": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "des": 12,
    "december": 12,
    "desember": 12,
}

MONTH_NAME_ID = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


@dataclass
class SchemaRow:
    no: int
    table_db: str
    output_stem: str
    source_url: str = ""


# Utility helpers
def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify(text: str) -> str:
    text = clean_text(text)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def strip_schema_name(table_db: str) -> str:
    table_db = clean_text(table_db)
    if "." in table_db:
        return table_db.split(".")[-1]
    return table_db


def output_name(schema: SchemaRow) -> str:
    return f"{schema.no}_{strip_schema_name(schema.table_db)}.csv"


def ensure_exists(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} tidak ditemukan: {path}")


def to_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    text = clean_text(value)
    if not text or text in {"-", "—"}:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


# Downloader: BI sources
def quarter_roman(q: int) -> str:
    return {1: "I", 2: "II", 3: "III", 4: "IV"}[q]


def quarter_page_code(q: int) -> str:
    return {1: "TWI", 2: "TWII", 3: "TWIII", 4: "TWIV"}[q]


def fetch_url_bytes(url: str, timeout: int = 60) -> bytes:
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


def download_url(url: str, dest: Path, force: bool = False, timeout: int = 90) -> bool:
    """Download URL to dest. Return True if downloaded, False if skipped/failed."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 0 and not force:
        print(f"[SKIP] Sudah ada: {dest}")
        return False

    tmp = dest.with_suffix(dest.suffix + ".part")
    try:
        data = fetch_url_bytes(url, timeout=timeout)
        if not data:
            raise ValueError("response kosong")
        tmp.write_bytes(data)
        tmp.replace(dest)
        print(f"[OK] Download: {url} -> {dest}")
        return True
    except (HTTPError, URLError, TimeoutError, ValueError, OSError) as exc:
        if tmp.exists():
            tmp.unlink(missing_ok=True)
        print(f"[WARN] Gagal download: {url} ({exc})", file=sys.stderr)
        return False


def safe_filename_from_url(url: str, fallback: str) -> str:
    parsed = urlparse(url)
    name = unquote(Path(parsed.path).name)
    name = clean_text(name)
    if not name or "." not in name:
        return fallback
    # Windows-safe enough.
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    return name


def extract_zip_links_from_suspi_html(html_text: str, page_url: str) -> List[str]:
    html_text = unescape(html_text)
    hrefs = re.findall(
        r"href=[\"']([^\"']+\.zip(?:\?[^\"']*)?)[\"']", html_text, flags=re.I
    )
    links: List[str] = []
    for href in hrefs:
        full = urljoin(page_url, href)
        low = unquote(full).lower()
        if "metadata" in low:
            continue
        if "suspi" not in low:
            continue
        if "tabel" not in low and "table" not in low:
            continue
        if full not in links:
            links.append(full)
    return links


def suspi_page_url(year: int, q: int) -> str:
    return f"{SUSPI_PAGES_URL}/SUSPI_{quarter_page_code(q)}_{year}.aspx"


def suspi_fallback_zip_urls(year: int, q: int) -> List[str]:
    roman = quarter_roman(q)
    quoted_older = f"Tabel%20SUSPI%20Triwulan%20{roman}%20-%20{year}.zip"
    quoted_older_no_space = f"Tabel%20SUSPI%20Triwulan%20{roman}-{year}.zip"
    return [
        f"{SUSPI_BASE_URL}/Document%20SUSPI/TABEL_SUSPI_TRIWULAN_{roman}_{year}.zip",
        f"{SUSPI_BASE_URL}/Documents/{quoted_older}",
        f"{SUSPI_BASE_URL}/Documents/{quoted_older_no_space}",
    ]


def download_suspi_quarter_zip(
    input_dir: Path, year: int, q: int, force: bool = False
) -> Optional[Path]:
    page_url = suspi_page_url(year, q)
    page_html = ""
    try:
        page_html = fetch_url_bytes(page_url, timeout=60).decode(
            "utf-8", errors="ignore"
        )
    except (HTTPError, URLError, TimeoutError, OSError) as exc:
        print(
            f"[WARN] Halaman SUSPI tidak bisa dibuka: {page_url} ({exc})",
            file=sys.stderr,
        )

    candidate_urls = []
    if page_html:
        candidate_urls.extend(extract_zip_links_from_suspi_html(page_html, page_url))
    candidate_urls.extend(
        [u for u in suspi_fallback_zip_urls(year, q) if u not in candidate_urls]
    )

    for url in candidate_urls:
        fallback = f"TABEL_SUSPI_TRIWULAN_{quarter_roman(q)}_{year}.zip"
        fname = safe_filename_from_url(url, fallback)
        dest = input_dir / "suspi_downloads" / fname
        if dest.exists() and dest.stat().st_size > 0 and not force:
            print(f"[SKIP] SUSPI {year} Q{q}: sudah ada -> {dest}")
            return dest
        ok = download_url(url, dest, force=force, timeout=120)
        if ok and dest.exists() and dest.stat().st_size > 0:
            return dest

    print(
        f"[WARN] Tidak berhasil download SUSPI {year} Q{q} dari halaman/fallback.",
        file=sys.stderr,
    )
    return None


def download_group_d_sources(
    input_dir: Path,
    start_year: int = 2019,
    end_year: int = 2025,
    force: bool = False,
    download_inflasi: bool = True,
    download_suspi: bool = True,
) -> None:
    input_dir.mkdir(parents=True, exist_ok=True)

    if download_inflasi:
        download_url(
            SEKI_TABEL8_1_URL, input_dir / "TABEL8_1.pdf", force=force, timeout=120
        )

    if download_suspi:
        print(
            f"[INFO] Download SUSPI source ZIP dari BI untuk {start_year}-{end_year}..."
        )
        for year in range(start_year, end_year + 1):
            for q in range(1, 5):
                download_suspi_quarter_zip(input_dir, year, q, force=force)
                time.sleep(0.2)  # sopan dikit ke server BI


# Schema reader
def read_group_d_schema(schema_path: Path) -> Dict[str, SchemaRow]:
    ensure_exists(schema_path, "Schema Excel")
    df = pd.read_excel(schema_path, sheet_name="Data Lengkap", engine="openpyxl")
    df.columns = [clean_text(c) for c in df.columns]

    group_col = "Grup"
    table_col = "Tabel DB"
    no_col = "No"
    source_col = "sumber URL/Dokumen"
    if (
        group_col not in df.columns
        or table_col not in df.columns
        or no_col not in df.columns
    ):
        raise ValueError(
            "Sheet 'Data Lengkap' harus punya kolom minimal: No, Tabel DB, Grup."
        )

    group_d = df[df[group_col].astype(str).str.strip().str.upper().eq("D")].copy()
    if group_d.empty:
        raise ValueError("Tidak ada baris Grup D di sheet 'Data Lengkap'.")

    result: Dict[str, SchemaRow] = {}
    for _, row in group_d.iterrows():
        table_db = clean_text(row.get(table_col))
        if not table_db or table_db.lower() == "nan":
            continue
        no = int(float(row.get(no_col)))
        src = clean_text(row.get(source_col, ""))
        schema_row = SchemaRow(
            no=no,
            table_db=table_db,
            output_stem=f"{no}_{strip_schema_name(table_db)}",
            source_url=src,
        )

        table_slug = strip_schema_name(table_db).lower()
        if "inflasi" in table_slug:
            result["inflasi"] = schema_row
        elif "hutang" in table_slug or "utang" in table_slug:
            result["hutang"] = schema_row

    missing = {"inflasi", "hutang"} - set(result)
    if missing:
        raise ValueError(
            f"Baris schema Grup D kurang lengkap. Tidak ketemu: {', '.join(sorted(missing))}"
        )
    return result


# Extractor 1: TABEL8_1.pdf -> Inflasi YoY
def extract_numbers_from_line(line: str) -> List[float]:
    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", line)
    return [float(x) for x in nums]


def month_tokens_from_line(line: str) -> List[str]:
    tokens = re.findall(r"[A-Za-z]+", line)
    out = []
    for tok in tokens:
        key = tok.lower()
        if key in MONTH_MAP:
            out.append(tok)
    return out


def assign_month_periods(
    start_year: int, month_tokens: Sequence[str]
) -> List[Tuple[int, int]]:
    periods: List[Tuple[int, int]] = []
    year = start_year
    last_month = None
    for tok in month_tokens:
        month = MONTH_MAP[tok.lower()]
        if last_month is not None and month < last_month:
            year += 1
        periods.append((year, month))
        last_month = month
    return periods


def build_page1_periods(text: str, n_values: int) -> List[Tuple[int, int, str]]:
    """
    Page Indonesia usually has annual columns 2020-2024, then monthly columns May-Aug.
    Annual columns are treated as December of that year, because SEKI annual columns are
    end-of-year snapshots in this table layout.
    """
    lines = text.splitlines()
    year_line = next(
        (ln for ln in lines[:8] if len(re.findall(r"20\d{2}", ln)) >= 2), ""
    )
    annual_years = [int(y) for y in re.findall(r"20\d{2}", year_line)]

    # keep the first run of years; for current PDFs it is 2020-2024.
    if len(annual_years) > 5:
        annual_years = annual_years[:5]

    month_line = ""
    for ln in lines[:12]:
        toks = month_tokens_from_line(ln)
        if toks:
            month_line = ln
            break
    months = month_tokens_from_line(month_line)

    periods: List[Tuple[int, int, str]] = [
        (y, 12, "annual_end_year") for y in annual_years
    ]
    if months:
        start_year = (annual_years[-1] + 1) if annual_years else 2025
        periods.extend(
            [(y, m, "monthly") for y, m in assign_month_periods(start_year, months)]
        )
    return periods[:n_values]


def build_page2_periods(text: str, n_values: int) -> List[Tuple[int, int, str]]:
    """Page English usually has monthly columns only, with years shown above month labels."""
    lines = text.splitlines()
    years: List[int] = []
    for ln in lines[:8]:
        found = [int(y) for y in re.findall(r"20\d{2}", ln)]
        if found:
            years.extend(found)
            if len(years) >= 1:
                break
    start_year = years[0] if years else 2025

    month_line = ""
    for ln in lines[:12]:
        toks = month_tokens_from_line(ln)
        if len(toks) >= 2:
            month_line = ln
            break
    months = month_tokens_from_line(month_line)
    return [(y, m, "monthly") for y, m in assign_month_periods(start_year, months)][
        :n_values
    ]


def extract_inflasi_yoy_from_pdf(pdf_path: Path, source_url: str = "") -> pd.DataFrame:
    if pdfplumber is None:
        raise ImportError(
            "pdfplumber belum terinstall. Install: pip install pdfplumber"
        )
    ensure_exists(pdf_path, "PDF TABEL8_1")

    records: List[dict] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            lines = [clean_text(ln) for ln in text.splitlines() if clean_text(ln)]

            yoy_line = ""
            for ln in lines:
                low = ln.lower()
                if "tahun ke tahun" in low or "year on year" in low:
                    yoy_line = ln
                    break
            if not yoy_line:
                continue

            values = extract_numbers_from_line(yoy_line)
            # Drop row number 39 if present at beginning/end.
            if values and values[0] == 39:
                values = values[1:]
            if values and values[-1] == 39:
                values = values[:-1]

            if not values:
                continue

            if "tahun ke tahun" in yoy_line.lower():
                periods = build_page1_periods(text, len(values))
            else:
                periods = build_page2_periods(text, len(values))

            # Fallback: if header inference fails, do not silently create wrong periods.
            if len(periods) != len(values):
                raise ValueError(
                    f"Gagal cocokkan header periode vs nilai inflasi di halaman {page_idx}. "
                    f"Periode={len(periods)}, nilai={len(values)}. Line: {yoy_line}"
                )

            for (year, month, period_source), val in zip(periods, values):
                records.append(
                    {
                        "periode": f"{year}-{month:02d}",
                        "tahun": year,
                        "bulan_angka": month,
                        "bulan": MONTH_NAME_ID[month],
                        "inflasi_yoy_persen": val,
                        "kategori": "UMUM",
                        "jenis_perubahan": "Tahun ke Tahun / Year on Year",
                        "periode_asal": period_source,
                        "satuan": "persen",
                        "data_source": source_url or str(pdf_path),
                    }
                )

    df = pd.DataFrame(records)
    if df.empty:
        raise ValueError(
            "Tidak ada baris inflasi YoY yang berhasil diekstrak dari PDF."
        )
    # Drop duplicates if bilingual pages overlap. Prefer monthly over annual_end_year if same period.
    df["_rank"] = df["periode_asal"].map({"monthly": 2, "annual_end_year": 1}).fillna(0)
    df = df.sort_values(["periode", "_rank"]).drop_duplicates("periode", keep="last")
    df = df.drop(columns=["_rank"]).sort_values("periode").reset_index(drop=True)
    return df


# Extractor 2: SUSPI IDR -> Total Public Sector sheet
def parse_quarter(value) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("*", "")
    text = re.sub(r"\s+", "", text).upper()

    m = re.search(r"(20\d{2})Q([1-4])", text)
    if m:
        return f"{m.group(1)}Q{m.group(2)}"
    m = re.search(r"Q([1-4])(20\d{2})", text)
    if m:
        return f"{m.group(2)}Q{m.group(1)}"
    return None


def quarter_year(q: str) -> int:
    return int(q[:4])


def quarter_num(q: str) -> int:
    return int(q[-1])


def find_total_public_sector_sheet(wb):
    for ws in wb.worksheets:
        if "total public sector" in ws.title.lower():
            return ws
    # fallback: use last sheet if title format changes but workbook still follows SUSPI template
    return wb.worksheets[-1]


def find_quarter_header_row(ws) -> Tuple[int, Dict[int, str]]:
    best_row = None
    best: Dict[int, str] = {}
    max_scan_rows = min(ws.max_row or 1, 15)
    for r in range(1, max_scan_rows + 1):
        found: Dict[int, str] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            q = parse_quarter(ws.cell(r, c).value)
            if q:
                found[c] = q
        if len(found) > len(best):
            best_row = r
            best = found
    if best_row is None or not best:
        raise ValueError(f"Tidak menemukan header kuartal di sheet {ws.title!r}")
    return best_row, best


def canonical_label(label: str) -> str:
    label = clean_text(label)
    # Remove common footnote markers: 1, 2), 1), etc.
    label = re.sub(r"\s*\d+\)?\s*$", "", label)
    label = label.replace(":", "")
    label = re.sub(r"\s+", " ", label).strip()
    return label


def is_section_label(label: str) -> bool:
    low = canonical_label(label).lower()
    if not low:
        return False
    return (
        low.startswith("by ")
        or low.startswith("memorandum item")
        or low.startswith("note")
        or low.startswith("validation")
        or low
        in {"gross general government debt", "total gross public sector debt position"}
    )


def is_subsection_label(label: str) -> bool:
    low = canonical_label(label).lower()
    return low.startswith("with payment due") or low.startswith("with payments due")


def indicator_slug(section: str, subsection: str, label: str) -> str:
    parts = [
        canonical_label(section),
        canonical_label(subsection),
        canonical_label(label),
    ]
    parts = [p for p in parts if p]
    return slugify("_".join(parts))


def extract_total_public_sector_from_workbook_bytes(
    content: bytes, source_name: str
) -> Tuple[pd.DataFrame, set]:
    if load_workbook is None:
        raise ImportError("openpyxl belum terinstall. Install: pip install openpyxl")

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    ws = find_total_public_sector_sheet(wb)
    header_row, quarter_cols = find_quarter_header_row(ws)

    # Keep only 2019-2025.
    quarter_cols = {
        c: q for c, q in quarter_cols.items() if 2019 <= quarter_year(q) <= 2025
    }
    if not quarter_cols:
        return pd.DataFrame(), set()

    current_section = ""
    current_subsection = ""
    records: List[dict] = []
    indicators_in_file = set()

    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        raw_label = ws.cell(r, 2).value
        label = clean_text(raw_label)
        if not label:
            continue

        label_clean = canonical_label(label)
        low = label_clean.lower()
        if low.startswith("note"):
            break

        if low == "total gross debt":
            # Extract total row, then use it as section for detail rows below it.
            ind = slugify(label_clean)
            indicators_in_file.add(ind)
            for c, q in quarter_cols.items():
                val = to_number(ws.cell(r, c).value)
                if val is not None:
                    records.append(
                        {
                            "periode": q,
                            "tahun": quarter_year(q),
                            "kuartal": f"Q{quarter_num(q)}",
                            "indicator": ind,
                            "value": val,
                            "source_file": source_name,
                            "sheet": ws.title,
                        }
                    )
            current_section = label_clean
            current_subsection = ""
            continue

        if is_section_label(label_clean):
            current_section = label_clean
            current_subsection = ""
            continue

        if is_subsection_label(label_clean):
            current_subsection = label_clean
            continue

        # Skip notes/validation/symbol rows.
        if (
            low.startswith("*)")
            or low.startswith("1.")
            or low.startswith("1 ")
            or low.startswith("-")
        ):
            continue

        # Extract numeric row only if at least one quarter cell has a number.
        values = {q: to_number(ws.cell(r, c).value) for c, q in quarter_cols.items()}
        values = {q: v for q, v in values.items() if v is not None}
        if not values:
            continue

        ind = indicator_slug(current_section, current_subsection, label_clean)
        if not ind:
            continue
        indicators_in_file.add(ind)
        for q, val in values.items():
            records.append(
                {
                    "periode": q,
                    "tahun": quarter_year(q),
                    "kuartal": f"Q{quarter_num(q)}",
                    "indicator": ind,
                    "value": val,
                    "source_file": source_name,
                    "sheet": ws.title,
                }
            )

    return pd.DataFrame(records), indicators_in_file


def iter_suspi_idr_workbooks(
    input_paths: Sequence[Path],
) -> Iterable[Tuple[str, bytes]]:
    """Yield (source_name, content_bytes) for every IDR xlsx in paths/zip/directories."""
    seen = set()
    for path in input_paths:
        if not path.exists():
            continue
        if path.is_dir():
            children = sorted([p for p in path.rglob("*") if p.is_file()])
            yield from iter_suspi_idr_workbooks(children)
            continue

        suffix = path.suffix.lower()
        name_lower = path.name.lower()
        if suffix == ".zip":
            with zipfile.ZipFile(path) as zf:
                for info in zf.infolist():
                    entry_lower = info.filename.lower()
                    if not entry_lower.endswith(".xlsx"):
                        continue
                    if "idr" not in entry_lower or "usd" in entry_lower:
                        continue
                    key = (str(path), info.filename)
                    if key in seen:
                        continue
                    seen.add(key)
                    yield f"{path.name}::{info.filename}", zf.read(info)
        elif suffix in {".xlsx", ".xlsm"}:
            # Direct xlsx: process if IDR is in filename, or if user explicitly puts only IDR files.
            if "usd" in name_lower:
                continue
            if "idr" in name_lower or "suspi" in name_lower:
                key = str(path)
                if key in seen:
                    continue
                seen.add(key)
                yield path.name, path.read_bytes()


def extract_suspi_total_public_sector(
    input_paths: Sequence[Path], source_url: str = ""
) -> pd.DataFrame:
    all_long: List[pd.DataFrame] = []
    workbooks_processed = []

    for source_name, content in iter_suspi_idr_workbooks(input_paths):
        try:
            df_long, _indicators = extract_total_public_sector_from_workbook_bytes(
                content, source_name
            )
        except Exception as exc:
            print(f"[WARN] Skip SUSPI workbook {source_name!r}: {exc}", file=sys.stderr)
            continue
        if df_long.empty:
            continue
        df_long["data_source"] = source_url or source_name
        all_long.append(df_long)
        workbooks_processed.append(source_name)

    if not all_long:
        raise ValueError("Tidak ada workbook SUSPI IDR yang berhasil diekstrak.")

    long_df = pd.concat(all_long, ignore_index=True)

    selected_indicators = {
        "total_gross_debt": "total_gross_debt",
        "by_residence_of_the_creditor_domestic_creditors": "domestic_creditors",
        "by_residence_of_the_creditor_external_creditors": "external_creditors",
        "by_currency_of_denomination_domestic_currency": "domestic_currency",
        "by_currency_of_denomination_foreign_currency": "foreign_currency",
    }

    long_df = long_df[long_df["indicator"].isin(selected_indicators.keys())].copy()
    if long_df.empty:
        raise ValueError(
            "Indikator hutang yang dipilih tidak ditemukan. "
            "Cek apakah workbook SUSPI IDR masih memakai sheet '4. Total Public Sector'."
        )

    # Jika periode yang sama muncul dari beberapa file, ambil dari file yang diproses terakhir.
    long_df["_source_order"] = long_df["source_file"].astype("category").cat.codes
    long_df = long_df.sort_values(["periode", "indicator", "_source_order"])
    long_df = long_df.drop_duplicates(["periode", "indicator"], keep="last")

    wide = long_df.pivot_table(
        index=["periode", "tahun", "kuartal"],
        columns="indicator",
        values="value",
        aggfunc="last",
    ).reset_index()
    wide.columns.name = None
    wide = wide.rename(columns=selected_indicators)

    # Urutan kolom final yang lebih manusiawi.
    final_cols = [
        "periode",
        "tahun",
        "kuartal",
        "total_gross_debt",
        "domestic_creditors",
        "external_creditors",
        "domestic_currency",
        "foreign_currency",
    ]
    for col in final_cols:
        if col not in wide.columns:
            wide[col] = pd.NA

    wide = wide[final_cols].sort_values("periode").reset_index(drop=True)

    # Biar CSV enak dibaca, angka dibulatkan 2 desimal.
    money_cols = [
        "total_gross_debt",
        "domestic_creditors",
        "external_creditors",
        "domestic_currency",
        "foreign_currency",
    ]
    for col in money_cols:
        wide[col] = pd.to_numeric(wide[col], errors="coerce").round(2)

    wide["satuan"] = "miliar IDR / billions of IDR"
    wide["data_source"] = source_url or "; ".join(workbooks_processed)
    return wide


# Main CLI
def find_inflasi_pdf(input_dir: Path, explicit_pdf: Optional[Path]) -> Path:
    if explicit_pdf:
        ensure_exists(explicit_pdf, "PDF inflasi")
        return explicit_pdf
    candidates = sorted(input_dir.rglob("TABEL8_1.pdf")) if input_dir.exists() else []
    if not candidates:
        candidates = sorted(input_dir.rglob("*.pdf")) if input_dir.exists() else []
        candidates = [
            p
            for p in candidates
            if "8_1" in p.name.lower() or "tabel8" in p.name.lower()
        ]
    if not candidates:
        raise FileNotFoundError(
            f"TABEL8_1.pdf tidak ditemukan di {input_dir}. Pakai --inflasi-pdf untuk path eksplisit."
        )
    return candidates[0]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract Grup D BI tables into clean CSV files."
    )
    parser.add_argument(
        "--input-dir",
        required=True,
        help="Folder sumber Grup D berisi TABEL8_1.pdf dan/atau SUSPI zip/xlsx.",
    )
    parser.add_argument(
        "--schema", required=True, help="Path Excel schema ekonomi (.xlsm/.xlsx)."
    )
    parser.add_argument("--output-dir", required=True, help="Folder output CSV.")
    parser.add_argument(
        "--inflasi-pdf", default=None, help="Optional: path eksplisit TABEL8_1.pdf."
    )
    parser.add_argument(
        "--suspi-path",
        action="append",
        default=None,
        help="Optional: path SUSPI zip/xlsx/folder. Bisa dipakai berkali-kali.",
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Matikan auto-download dari web BI; hanya pakai file lokal.",
    )
    parser.add_argument(
        "--force-download",
        action="store_true",
        help="Download ulang walaupun file sudah ada.",
    )
    parser.add_argument(
        "--suspi-start-year",
        type=int,
        default=2019,
        help="Tahun awal SUSPI yang di-download/diekstrak. Default 2019.",
    )
    parser.add_argument(
        "--suspi-end-year",
        type=int,
        default=2025,
        help="Tahun akhir SUSPI yang di-download/diekstrak. Default 2025.",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    schema_path = Path(args.schema)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    schemas = read_group_d_schema(schema_path)

    if args.suspi_start_year > args.suspi_end_year:
        raise ValueError(
            "--suspi-start-year tidak boleh lebih besar dari --suspi-end-year"
        )

    if not args.no_download:
        download_group_d_sources(
            input_dir=input_dir,
            start_year=args.suspi_start_year,
            end_year=args.suspi_end_year,
            force=args.force_download,
            download_inflasi=(args.inflasi_pdf is None),
            download_suspi=(args.suspi_path is None),
        )

    # 1) Inflasi - TABEL8_1.pdf
    inflasi_pdf = find_inflasi_pdf(
        input_dir, Path(args.inflasi_pdf) if args.inflasi_pdf else None
    )
    inflasi_df = extract_inflasi_yoy_from_pdf(
        inflasi_pdf, schemas["inflasi"].source_url
    )
    inflasi_out = output_dir / output_name(schemas["inflasi"])
    inflasi_df.to_csv(inflasi_out, index=False, encoding="utf-8-sig")
    print(f"[OK] Inflasi: {len(inflasi_df):,} rows -> {inflasi_out}")

    # 2) Hutang - SUSPI IDR zip/xlsx, Total Public Sector only
    suspi_paths = [Path(p) for p in args.suspi_path] if args.suspi_path else [input_dir]
    hutang_df = extract_suspi_total_public_sector(
        suspi_paths, schemas["hutang"].source_url
    )
    hutang_out = output_dir / output_name(schemas["hutang"])
    hutang_df.to_csv(hutang_out, index=False, encoding="utf-8-sig")
    print(
        f"[OK] Hutang SUSPI: {len(hutang_df):,} rows, {len(hutang_df.columns):,} cols -> {hutang_out}"
    )

    print("\nSelesai. Output mengikuti nomor + nama tabel DB dari sheet schema Grup D.")


if __name__ == "__main__":
    main()
